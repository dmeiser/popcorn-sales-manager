"""Unit tests for report generation Lambda handler.

Updated for multi-table design (campaigns, orders tables).
"""

import os
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict

import boto3
import openpyxl
import pytest

from src.handlers.report_generation import request_campaign_report


def get_orders_table() -> Any:
    """Get orders table for testing."""
    dynamodb = boto3.resource("dynamodb", region_name="us-east-1")
    return dynamodb.Table("kernelworx-orders-v2-ue1-dev")


@pytest.fixture
def sample_orders(dynamodb_table: Any, sample_profile_id: str, sample_campaign_id: str) -> list[Dict[str, Any]]:
    """Sample order data for testing (multi-table design)."""
    orders = [
        {
            "orderId": "ORDER#order-1",
            "campaignId": sample_campaign_id,  # DynamoDB schema uses campaignId as PK
            "profileId": sample_profile_id,
            "customerName": "John Doe",
            "customerPhone": "555-123-4567",
            "customerAddress": {
                "street": "123 Main St",
                "city": "Anytown",
                "state": "CA",
                "zipCode": "12345",
            },
            "paymentMethod": "CASH",
            "totalAmount": Decimal("45.00"),
            "lineItems": [
                {"productName": "Product A", "quantity": 2, "price": Decimal("10.00")},
                {"productName": "Product B", "quantity": 1, "price": Decimal("25.00")},
            ],
            "createdAt": "2025-09-15T10:00:00+00:00",
        },
        {
            "orderId": "ORDER#order-2",
            "campaignId": sample_campaign_id,  # DynamoDB schema uses campaignId as PK
            "profileId": sample_profile_id,
            "customerName": "Jane Smith",
            "customerEmail": "jane@example.com",
            "paymentMethod": "CHECK",
            "totalAmount": Decimal("30.00"),
            "lineItems": [
                {"productName": "Product C", "quantity": 3, "price": Decimal("10.00")},
            ],
            "createdAt": "2025-09-20T14:30:00+00:00",
        },
    ]

    # Insert into orders table (not profiles table)
    orders_table = get_orders_table()
    for order in orders:
        orders_table.put_item(Item=order)

    return orders


class TestRequestCampaignReport:
    """Tests for requestCampaignReport Lambda handler."""

    def test_owner_can_generate_excel_report(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_account_id: str,
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that profile owner can generate Excel report."""
        # Arrange
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert "reportId" in result
        assert result["reportId"].startswith("REPORT#")
        assert "reportUrl" in result
        assert result["status"] == "COMPLETED"
        assert "expiresAt" in result

        # Verify S3 upload
        bucket_name = os.environ.get("EXPORTS_BUCKET", "test-exports-bucket")
        objects = s3_bucket.list_objects_v2(Bucket=bucket_name)
        assert objects["KeyCount"] == 1
        assert objects["Contents"][0]["Key"].endswith(".xlsx")

        # Verify Excel content
        obj = s3_bucket.get_object(Bucket=bucket_name, Key=objects["Contents"][0]["Key"])
        wb = openpyxl.load_workbook(BytesIO(obj["Body"].read()))
        ws = wb.active

        # Check headers (row 1 - new format without title row)
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=1, column=2).value == "Phone"
        assert ws.cell(row=1, column=3).value == "Address"

        # Check data rows (starting at row 2)
        assert ws.cell(row=2, column=1).value == "John Doe"
        assert ws.cell(row=3, column=1).value == "Jane Smith"

    def test_owner_can_generate_csv_report(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that profile owner can generate CSV report."""
        # Arrange
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "csv"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert result["status"] == "COMPLETED"
        assert "reportUrl" in result

        # Verify S3 upload
        bucket_name = os.environ.get("EXPORTS_BUCKET", "test-exports-bucket")
        objects = s3_bucket.list_objects_v2(Bucket=bucket_name)
        assert objects["KeyCount"] == 1
        assert objects["Contents"][0]["Key"].endswith(".csv")

        # Verify CSV content
        obj = s3_bucket.get_object(Bucket=bucket_name, Key=objects["Contents"][0]["Key"])
        csv_content = obj["Body"].read().decode("utf-8")

        # Check headers and data
        assert "Name,Phone,Address" in csv_content
        assert "John Doe" in csv_content
        assert "Jane Smith" in csv_content

    def test_contributor_with_read_can_generate_report(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_profile_id: str,
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
        another_account_id: str,
        shares_table: Any,
    ) -> None:
        """Test that contributor with READ permission can generate report."""
        # Create share with READ permission (now in dedicated shares table)
        # targetAccountId must have ACCOUNT# prefix to match auth.py lookup
        shares_table.put_item(
            Item={
                "profileId": sample_profile_id,
                "targetAccountId": f"ACCOUNT#{another_account_id}",
                "permissions": ["READ"],
                "grantedAt": datetime.now(timezone.utc).isoformat(),
            }
        )

        event = {
            **appsync_event,
            "identity": {"sub": another_account_id},
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert result["status"] == "COMPLETED"

    def test_non_owner_without_share_cannot_generate_report(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
        another_account_id: str,
    ) -> None:
        """Test that non-owner without share cannot generate report."""
        event = {
            **appsync_event,
            "identity": {"sub": another_account_id},
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert "errorCode" in result
        assert result["errorCode"] == "FORBIDDEN"

    def test_nonexistent_campaign_returns_error(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that requesting report for non-existent campaign returns error."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": "CAMPAIGN#nonexistent", "format": "xlsx"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert "errorCode" in result
        assert result["errorCode"] == "NOT_FOUND"

    def test_default_format_is_xlsx(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that default format is xlsx when not specified."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id}},  # No format specified
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert result["status"] == "COMPLETED"

        # Verify S3 upload is xlsx
        bucket_name = os.environ.get("EXPORTS_BUCKET", "test-exports-bucket")
        objects = s3_bucket.list_objects_v2(Bucket=bucket_name)
        assert objects["Contents"][0]["Key"].endswith(".xlsx")

    def test_empty_campaign_generates_report_with_no_orders(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that empty campaign (no orders) generates valid report."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "csv"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert
        assert result["status"] == "COMPLETED"

        # Verify CSV content is just headers (no data rows)
        bucket_name = os.environ.get("EXPORTS_BUCKET", "test-exports-bucket")
        objects = s3_bucket.list_objects_v2(Bucket=bucket_name)
        obj = s3_bucket.get_object(Bucket=bucket_name, Key=objects["Contents"][0]["Key"])
        csv_content = obj["Body"].read().decode("utf-8")

        # Should have headers but no data rows (just Name,Phone,Address,Total)
        assert "Name,Phone,Address,Total" in csv_content
        lines = csv_content.strip().split("\n")
        assert len(lines) == 1  # Only header row

    def test_presigned_url_expiration_is_7_days(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test that pre-signed URL expires in 7 days."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # Act
        result = request_campaign_report(event, lambda_context)

        # Assert - verify expiresAt is approximately 7 days from now
        expires_at = datetime.fromisoformat(result["expiresAt"].replace("Z", "+00:00"))
        now = datetime.now(timezone.utc)
        time_diff = (expires_at - now).total_seconds()

        # Should be approximately 7 days (604800 seconds)
        # Allow 60 second tolerance for test execution time
        assert 604740 <= time_diff <= 604860

    def test_generic_exception_returns_internal_error(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
        monkeypatch: Any,
    ) -> None:
        """Test that generic exception returns INTERNAL_ERROR."""
        from unittest.mock import patch

        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # Mock _get_campaign to raise an unexpected exception
        with patch(
            "src.handlers.report_generation._get_campaign",
            side_effect=ValueError("Unexpected error"),
        ):
            result = request_campaign_report(event, lambda_context)

        # Assert
        assert "errorCode" in result
        assert result["errorCode"] == "INTERNAL_ERROR"
        assert "Unexpected error" in result["message"]

    def test_excel_cell_with_problematic_value(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test Excel generation handles cells with problematic values gracefully."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # This test indirectly exercises the except clause at line 290
        # by ensuring Excel generation completes successfully even with
        # various cell value types and edge cases in the actual data
        result = request_campaign_report(event, lambda_context)

        assert "reportId" in result
        assert result["status"] == "COMPLETED"

    def test_excel_with_merged_cells(
        self,
        dynamodb_table: Any,
        s3_bucket: Any,
        sample_profile: Dict[str, Any],
        sample_campaign: Dict[str, Any],
        sample_orders: list[Dict[str, Any]],
        sample_campaign_id: str,
        appsync_event: Dict[str, Any],
        lambda_context: Any,
    ) -> None:
        """Test Excel generation handles merged cells that may lack column_letter attribute."""
        event = {
            **appsync_event,
            "arguments": {"input": {"campaignId": sample_campaign_id, "format": "xlsx"}},
        }

        # This test exercises the column_letter is None branch (line 285 continue)
        # in the auto-sizing columns logic for handling edge cases like MergedCells
        result = request_campaign_report(event, lambda_context)

        assert "reportId" in result
        assert result["status"] == "COMPLETED"
        assert "reportUrl" in result


class TestFormatAddress:
    """Tests for the _format_address helper function."""

    def test_empty_address(self) -> None:
        """Test formatting empty address returns empty string."""
        from src.handlers.report_generation import _format_address

        assert _format_address(None) == ""
        assert _format_address({}) == ""

    def test_full_address(self) -> None:
        """Test formatting complete address."""
        from src.handlers.report_generation import _format_address

        address = {
            "street": "123 Main St",
            "city": "Anytown",
            "state": "CA",
            "zipCode": "12345",
        }
        result = _format_address(address)
        assert result == "123 Main St, Anytown CA 12345"

    def test_address_with_no_street(self) -> None:
        """Test address with only city/state/zip (no street)."""
        from src.handlers.report_generation import _format_address

        address = {
            "city": "Anytown",
            "state": "CA",
            "zipCode": "12345",
        }
        result = _format_address(address)
        assert result == "Anytown CA 12345"

    def test_address_with_only_street(self) -> None:
        """Test address with only street (no city/state/zip)."""
        from src.handlers.report_generation import _format_address

        address = {
            "street": "123 Main St",
        }
        result = _format_address(address)
        assert result == "123 Main St"

    def test_address_with_empty_city_state_zip(self) -> None:
        """Test address with city/state/zip all empty strings."""
        from src.handlers.report_generation import _format_address

        # This tests the edge case where city, state, zipCode exist but are empty
        address = {
            "street": "123 Main St",
            "city": "",
            "state": "",
            "zipCode": "",
        }
        result = _format_address(address)
        # city_state_zip will be empty string after filter and join
        assert result == "123 Main St"

    def test_address_partial_city_state(self) -> None:
        """Test address with only some city/state/zip fields."""
        from src.handlers.report_generation import _format_address

        # Only city and state
        address = {"city": "Anytown", "state": "CA"}
        result = _format_address(address)
        assert result == "Anytown CA"

        # Only zipCode
        address = {"zipCode": "12345"}
        result = _format_address(address)
        assert result == "12345"
