"""
Report generation Lambda handler.

Implements:
- requestCampaignReport: Generate Excel/CSV report for campaign data
"""

import os
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import TYPE_CHECKING, Any, Dict

import boto3

if TYPE_CHECKING:  # pragma: no cover
    from mypy_boto3_s3.client import S3Client
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

# Handle both Lambda (absolute) and unit test (relative) imports
try:  # pragma: no cover
    from utils.auth import check_profile_access
    from utils.dynamodb import get_required_env, tables
    from utils.errors import AppError, ErrorCode
    from utils.logging import get_logger
except ModuleNotFoundError:  # pragma: no cover
    from ..utils.auth import check_profile_access
    from ..utils.dynamodb import get_required_env, tables
    from ..utils.errors import AppError, ErrorCode
    from ..utils.logging import get_logger


# Module-level proxy that tests can monkeypatch
s3_client: "S3Client | None" = None


def _get_s3_client() -> "S3Client":
    """Return the S3 client (module-level override for tests, otherwise a fresh boto3 client)."""
    global s3_client
    if s3_client is not None:
        return s3_client
    return boto3.client("s3", endpoint_url=os.getenv("S3_ENDPOINT"))


def request_campaign_report(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Generate a campaign report and upload to S3.

    GraphQL mutation: requestCampaignReport(input: RequestCampaignReportInput!)

    Returns:
        {
          reportId: String!
          campaignId: String!
          profileId: String!
          reportUrl: String
          status: String!
          createdAt: String!
          expiresAt: String
        }
    """
    logger = get_logger(__name__, event.get("requestId", "unknown"))

    try:
        # Extract arguments - GraphQL passes campaignId, but we store as campaignId in DynamoDB
        args = event["arguments"]["input"]
        campaign_id = args["campaignId"]  # GraphQL uses campaignId
        campaign_id = campaign_id  # Map to internal campaignId for DynamoDB queries
        report_format = args.get("format", "xlsx")  # xlsx or csv
        caller_account_id = event["identity"]["sub"]

        logger.info(
            "Generating campaign report",
            campaign_id=campaign_id,
            format=report_format,
            caller_account_id=caller_account_id,
        )

        # Get campaign and verify authorization (multi-table design)
        campaign = _get_campaign(tables.campaigns, campaign_id)

        if not campaign:
            raise AppError(ErrorCode.NOT_FOUND, f"Campaign {campaign_id} not found")

        profile_id = campaign["profileId"]

        # Check authorization (must have read access to profile)
        if not check_profile_access(caller_account_id, profile_id, "read"):
            raise AppError(ErrorCode.FORBIDDEN, "You don't have access to this campaign")

        # Get all orders for the campaign (multi-table design)
        orders = _get_campaign_orders(tables.orders, campaign_id)

        # Generate report
        if report_format.lower() == "csv":
            report_content = _generate_csv_report(campaign, orders)
            content_type = "text/csv"
            file_extension = "csv"
        else:
            report_content = _generate_excel_report(campaign, orders)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_extension = "xlsx"

        # Upload to S3
        report_id = f"REPORT#{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
        exports_bucket = get_required_env("EXPORTS_BUCKET")
        s3_key = f"reports/{profile_id}/{campaign_id}/{report_id}.{file_extension}"

        s3 = _get_s3_client()
        s3.put_object(
            Bucket=exports_bucket,
            Key=s3_key,
            Body=report_content,
            ContentType=content_type,
        )

        # Generate pre-signed URL (valid for 7 days)
        expiration = 7 * 24 * 60 * 60  # 7 days in seconds
        report_url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": exports_bucket, "Key": s3_key},
            ExpiresIn=expiration,
        )

        now = datetime.now(timezone.utc)
        expires_at = now + timedelta(days=7)

        result = {
            "reportId": report_id,
            "campaignId": campaign_id,  # Return campaignId for GraphQL API
            "profileId": profile_id,
            "reportUrl": report_url,
            "status": "COMPLETED",
            "createdAt": now.isoformat(),
            "expiresAt": expires_at.isoformat(),
        }

        logger.info("Report generated successfully", report_id=report_id, s3_key=s3_key)
        return result

    except AppError as e:
        logger.error(f"AppError: {e.error_code}: {e.message}")
        return e.to_dict()  # type: ignore[no-any-return]
    except Exception as e:
        logger.error("Unexpected error generating report", error=str(e), exc_info=True)
        error = AppError(ErrorCode.INTERNAL_ERROR, f"Failed to generate report: {str(e)}")
        return error.to_dict()  # type: ignore[no-any-return]


def _get_campaign(table: Any, campaign_id: str) -> Dict[str, Any] | None:
    """Get campaign by ID (V2: Query campaignId-index GSI since PK=profileId, SK=campaignId)."""
    # Campaign ID format: CAMPAIGN#uuid
    # V2: Campaigns are stored with PK=profileId, SK=campaignId
    # Use campaignId-index GSI for lookup by campaignId
    response = table.query(
        IndexName="campaignId-index",
        KeyConditionExpression="campaignId = :campaignId",
        ExpressionAttributeValues={":campaignId": campaign_id},
        Limit=1,
    )
    items = response.get("Items", [])
    item: Dict[str, Any] | None = items[0] if items else None
    return item


def _get_campaign_orders(table: Any, campaign_id: str) -> list[Dict[str, Any]]:
    """Get all orders for a campaign (V2: Direct PK query since PK=campaignId)."""
    # V2 schema: Orders table has PK=campaignId, SK=orderId
    # No GSI needed - direct query on the partition key
    response = table.query(
        KeyConditionExpression="campaignId = :campaign_id",
        ExpressionAttributeValues={
            ":campaign_id": campaign_id,
        },
    )

    items = response.get("Items", [])
    return list(items) if items else []


def _format_address(address: Dict[str, Any] | None) -> str:
    """Format address object as string."""
    if not address:
        return ""
    parts = []
    if address.get("street"):
        parts.append(address["street"])
    if address.get("city") or address.get("state") or address.get("zipCode"):
        city_state_zip = " ".join(
            filter(
                None,
                [address.get("city"), address.get("state"), address.get("zipCode")],
            )
        )
        if city_state_zip:  # pragma: no branch - always true if we entered outer if
            parts.append(city_state_zip)
    return ", ".join(parts)


def _get_unique_products(orders: list[Dict[str, Any]]) -> list[str]:
    """Get sorted list of unique product names from orders."""
    return sorted(
        set(
            item.get("productName", "")
            for order in orders
            for item in order.get("lineItems", [])
            if item.get("productName")
        )
    )


def _get_product_quantities(order: Dict[str, Any]) -> dict[str, int]:
    """Get product quantities for an order, summing duplicates."""
    quantities: dict[str, int] = {}
    for item in order.get("lineItems", []):
        product_name = item.get("productName", "")
        quantity = item.get("quantity", 0)
        quantities[product_name] = quantities.get(product_name, 0) + quantity
    return quantities


def _generate_csv_report(campaign: Dict[str, Any], orders: list[Dict[str, Any]]) -> bytes:
    """Generate CSV report with product columns."""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    all_products = _get_unique_products(orders)
    headers = ["Name", "Phone", "Address"] + all_products + ["Total"]
    writer.writerow(headers)

    for order in orders:
        row = [
            order.get("customerName", ""),
            order.get("customerPhone", ""),
            _format_address(order.get("customerAddress", {})),
        ]
        quantities = _get_product_quantities(order)
        for product in all_products:
            row.append(quantities.get(product, ""))
        row.append(order.get("totalAmount", 0))
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def _write_excel_headers(ws: Any, headers: list[str]) -> None:
    """Write styled headers to Excel worksheet."""
    from openpyxl.styles import Font, PatternFill
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font


def _write_excel_order_row(
    ws: Any, row_idx: int, order: Dict[str, Any], all_products: list[str], headers: list[str]
) -> None:
    """Write a single order row to Excel worksheet."""
    ws.cell(row=row_idx, column=1, value=order.get("customerName", ""))
    ws.cell(row=row_idx, column=2, value=order.get("customerPhone", ""))
    ws.cell(row=row_idx, column=3, value=_format_address(order.get("customerAddress", {})))

    quantities = _get_product_quantities(order)
    for col_idx, product in enumerate(all_products, start=4):
        ws.cell(row=row_idx, column=col_idx, value=quantities.get(product, ""))
    ws.cell(row=row_idx, column=len(headers), value=float(order.get("totalAmount", 0)))


def _autosize_excel_columns(ws: Any) -> None:
    """Auto-size columns in Excel worksheet."""
    for column in ws.columns:
        first_cell = column[0]
        column_letter = getattr(first_cell, "column_letter", None)
        if column_letter is None:  # pragma: no cover
            continue  # pragma: no cover
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _generate_excel_report(campaign: Dict[str, Any], orders: list[Dict[str, Any]]) -> bytes:
    """Generate Excel report with product columns."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook must have an active worksheet"
    ws.title = "Orders"

    all_products = _get_unique_products(orders)
    headers = ["Name", "Phone", "Address"] + all_products + ["Total"]

    _write_excel_headers(ws, headers)
    for row_idx, order in enumerate(orders, start=2):
        _write_excel_order_row(ws, row_idx, order, all_products, headers)
    _autosize_excel_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
